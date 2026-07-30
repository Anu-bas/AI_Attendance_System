"""
Generate downloadable attendance reports in PDF, Excel and CSV formats.
"""
import os
import csv
import io
from datetime import datetime

import matplotlib
matplotlib.use("Agg")  # headless backend, safe for server-side rendering
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from config import Config

VIOLET = "6C5CE7"
BLUE = "3B5BFE"


def _timestamped_name(prefix, ext):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def generate_csv_report(records, summary):
    """records: list of dicts (attendance rows). Returns path to saved CSV."""
    filename = _timestamped_name("attendance_report", "csv")
    path = os.path.join(Config.REPORTS_DIR, filename)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Attendance Report Summary"])
        for k, v in summary.items():
            writer.writerow([k, v])
        writer.writerow([])
        writer.writerow(["Student ID", "Name", "Department", "Year", "Section",
                         "Date", "Time", "Subject", "Faculty", "Status", "Confidence(%)"])
        for r in records:
            writer.writerow([
                r.get("register_number"), r.get("name"), r.get("department"),
                r.get("year"), r.get("section"), r.get("attendance_date"),
                r.get("attendance_time"), r.get("subject"), r.get("faculty_name"),
                r.get("status"), r.get("confidence"),
            ])
    return path


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def generate_excel_report(records, summary):
    filename = _timestamped_name("attendance_report", "xlsx")
    path = os.path.join(Config.REPORTS_DIR, filename)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws_summary["A1"] = "AI Face Detection Attendance System - Report Summary"
    ws_summary["A1"].font = Font(bold=True, size=14, color=BLUE)
    ws_summary.merge_cells("A1:B1")

    row = 3
    for k, v in summary.items():
        ws_summary.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=v)
        row += 1
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 20

    ws = wb.create_sheet("Attendance Records")
    headers = ["Student ID", "Name", "Department", "Year", "Section",
               "Date", "Time", "Subject", "Faculty", "Status", "Confidence(%)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r_idx, r in enumerate(records, start=2):
        values = [
            r.get("register_number"), r.get("name"), r.get("department"),
            r.get("year"), r.get("section"), str(r.get("attendance_date")),
            str(r.get("attendance_time")), r.get("subject"), r.get("faculty_name"),
            r.get("status"), r.get("confidence"),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r.get("status") == "Present":
                cell.fill = PatternFill(start_color="E8F8EF", end_color="E8F8EF", fill_type="solid")
            elif r.get("status") == "Absent":
                cell.fill = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Chart image helpers (used inside the PDF, and reusable for dashboard export)
# ---------------------------------------------------------------------------
def _save_pie_chart(present, absent, out_path):
    plt.figure(figsize=(4, 4))
    values = [present, absent]
    labels = [f"Present ({present})", f"Absent ({absent})"]
    colors_ = ["#6C5CE7", "#FF7675"]
    plt.pie(values, labels=labels, colors=colors_, autopct="%1.1f%%", startangle=90)
    plt.title("Present vs Absent")
    plt.tight_layout()
    plt.savefig(out_path, dpi=140)
    plt.close()


def _save_bar_chart(student_names, percentages, out_path):
    plt.figure(figsize=(6, 4))
    plt.bar(student_names, percentages, color="#3B5BFE")
    plt.ylabel("Attendance %")
    plt.title("Student-wise Attendance Percentage")
    plt.xticks(rotation=45, ha="right", fontsize=7)
    plt.tight_layout()
    plt.savefig(out_path, dpi=140)
    plt.close()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def generate_pdf_report(records, summary, student_stats=None):
    """
    records: attendance rows (dicts)
    summary: dict of summary stats
    student_stats: list of {name, percentage} for the bar chart (optional)
    """
    filename = _timestamped_name("attendance_report", "pdf")
    path = os.path.join(Config.REPORTS_DIR, filename)

    doc = SimpleDocTemplate(path, pagesize=A4,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleViolet", parent=styles["Title"],
                                  textColor=colors.HexColor("#3B1F91"))
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"],
                                   textColor=colors.HexColor("#6C5CE7"))

    elements = []
    elements.append(Paragraph("AI Face Detection Attendance System", title_style))
    elements.append(Paragraph("Attendance Report", styles["Heading3"]))
    elements.append(Paragraph(datetime.now().strftime("Generated on %d %b %Y, %I:%M %p"),
                               styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Summary table
    elements.append(Paragraph("Summary", heading_style))
    summary_data = [[k, str(v)] for k, v in summary.items()]
    summary_table = Table(summary_data, colWidths=[7 * cm, 7 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EDE9FE")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 16))

    # Pie chart
    if "Present" in summary or True:
        present = summary.get("Present Students", 0)
        absent = summary.get("Absent Students", 0)
        pie_path = os.path.join(Config.REPORTS_DIR, "_tmp_pie.png")
        _save_pie_chart(present, absent, pie_path)
        elements.append(Paragraph("Present vs Absent", heading_style))
        elements.append(RLImage(pie_path, width=8 * cm, height=8 * cm))
        elements.append(Spacer(1, 12))

    # Bar chart of student-wise attendance %
    if student_stats:
        bar_path = os.path.join(Config.REPORTS_DIR, "_tmp_bar.png")
        names = [s["name"] for s in student_stats][:15]
        pct = [s["percentage"] for s in student_stats][:15]
        _save_bar_chart(names, pct, bar_path)
        elements.append(Paragraph("Student-wise Attendance %", heading_style))
        elements.append(RLImage(bar_path, width=16 * cm, height=9 * cm))
        elements.append(Spacer(1, 12))

    # Records table
    elements.append(Paragraph("Attendance Records", heading_style))
    table_header = ["ID", "Name", "Dept", "Date", "Time", "Subject", "Status"]
    table_data = [table_header]
    for r in records[:500]:  # cap to keep PDF manageable
        table_data.append([
            r.get("register_number"), r.get("name"), r.get("department"),
            str(r.get("attendance_date")), str(r.get("attendance_time")),
            r.get("subject"), r.get("status"),
        ])

    rec_table = Table(table_data, repeatRows=1)
    rec_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6C5CE7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F3FF")]),
    ]))
    elements.append(rec_table)

    doc.build(elements)

    for tmp in ("_tmp_pie.png", "_tmp_bar.png"):
        p = os.path.join(Config.REPORTS_DIR, tmp)
        if os.path.exists(p):
            os.remove(p)

    return path
